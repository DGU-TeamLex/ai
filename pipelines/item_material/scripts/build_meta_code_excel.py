#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
meta_code_glossary.csv 를 엑셀로 변환.
시트 3개(원재료코드/원재료리스크코드/수요리스크코드)로 나누고,
각 코드가 실제로 몇 건/얼마의 사용량에 걸려있는지 통계를 같이 붙인다.
"""
import csv
import sys
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

if len(sys.argv) != 4:
    raise SystemExit(
        "usage: build_meta_code_excel.py <glossary.csv> <mapping.csv> <output.xlsx>"
    )

GLOSSARY_FILE = sys.argv[1]
DATA_FILE = sys.argv[2]
OUT_FILE = sys.argv[3]

CATEGORY_SHEET = [
    ("raw_material", "원재료코드", "raw_material_meta_code"),
    ("raw_material_risk", "원재료리스크코드", "raw_material_risk_meta_code"),
    ("demand_risk", "수요리스크코드", "demand_risk_meta_code"),
]

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def main():
    with open(GLOSSARY_FILE, encoding="utf-8-sig") as f:
        glossary_rows = list(csv.DictReader(f))

    with open(DATA_FILE, encoding="utf-8-sig") as f:
        data_rows = list(csv.DictReader(f))

    wb = Workbook()
    wb.remove(wb.active)

    def _num(row, key):
        try:
            return float(row.get(key) or 0)
        except (TypeError, ValueError):
            return 0.0

    for cat, sheet_title, col in CATEGORY_SHEET:
        code_count = Counter()
        code_occurrence = Counter()
        code_usage = Counter()
        for r in data_rows:
            val = r.get(col, "")
            if not val:
                continue
            for code in set(val.split(";")):
                code_count[code] += 1
                code_occurrence[code] += _num(r, "occurrence_count")
                code_usage[code] += _num(r, "usage_sum")

        ws = wb.create_sheet(title=sheet_title)
        has_stage = cat == "raw_material"
        if has_stage:
            headers = [
                "메타코드", "설명", "공급단계", "단계신뢰도", "단계 비고",
                "매핑 품목건수", "누적 발주흔적", "누적 사용량(참고·결손多)"
            ]
        else:
            headers = [
                "메타코드", "설명", "매핑 품목건수", "누적 발주흔적",
                "누적 사용량(참고·결손多)"
            ]
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER

        cat_rows = []
        seen = set()
        for glossary_row in glossary_rows:
            code = glossary_row["meta_code"]
            if glossary_row["category"] == cat and code not in seen:
                seen.add(code)
                cat_rows.append(glossary_row)
        cat_rows.sort(key=lambda g: -code_occurrence.get(g["meta_code"], 0))

        for g in cat_rows:
            code = g["meta_code"]
            if has_stage:
                ws.append([
                    code,
                    g["description"],
                    g.get("supply_stage", ""),
                    g.get("stage_confidence", ""),
                    g.get("supply_stage_note", ""),
                    code_count.get(code, 0),
                    code_occurrence.get(code, 0),
                    code_usage.get(code, 0),
                ])
            else:
                ws.append([
                    code,
                    g["description"],
                    code_count.get(code, 0),
                    code_occurrence.get(code, 0),
                    code_usage.get(code, 0),
                ])

        last_col = len(headers)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=last_col):
            for cell in row:
                cell.border = BORDER
            for column_index in range(last_col - 3, last_col):
                row[column_index].alignment = Alignment(horizontal="right")
                row[column_index].number_format = "#,##0"

        widths = [36, 50, 15, 12, 34, 14, 14, 18] if has_stage else [40, 55, 14, 14, 18]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"

    wb.save(OUT_FILE)
    print("저장:", OUT_FILE)
    for cat, sheet_title, col in CATEGORY_SHEET:
        n = len([g for g in glossary_rows if g["category"] == cat])
        print(f"  - {sheet_title}: {n}개 코드")


if __name__ == "__main__":
    main()
